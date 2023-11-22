from datetime import timedelta

from django.contrib.auth.models import User
from django.http import JsonResponse
from django.shortcuts import render, redirect
from .models import Classes, Niveau, TypeAbonnements, CustomUser
from .forms import TypeAbonnementsForm, CustomUserForm, CustomLoginForm, LoginForm, RegisterForm
from django.contrib import messages
from django.contrib.auth.views import LoginView
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.hashers import make_password
from django.utils import timezone
import xlrd
from openpyxl import load_workbook


# Create your views here.

class CustomLoginView(LoginView):
    template_name = 'login.html'
    authentication_form = CustomLoginForm

    def form_valid(self, form):
        # Connexion de l'utilisateur
        user = form.get_user()
        print(user)
        login(user)

        # Vérifie si l'utilisateur est le superutilisateur et a choisi de créer un autre utilisateur
        if user.is_superuser and form.cleaned_data.get('create_user'):
            # Redirection vers la création d'utilisateur
            return redirect('dashboard')  # Remplacez 'creer_utilisateur' par l'URL de création d'utilisateur
        else:
            # Redirection vers l'application
            return redirect('dashboard')  # Remplacez 'votre_application' par l'URL de l'application

def login_page(request):
    if request.method == 'POST':
        username =request.POST.get("username")
        password = request.POST.get("password")
        create_user = request.POST.get("newuser")
        print(create_user)
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            message = f'Bonjour, {user.username}! Vous êtes connecté.'
            if create_user is not None and user.is_superuser:
                return redirect('register_page')
            return redirect('dashboard')
        else:
            message = 'Identifiants invalides.'
            return render(request, 'login.html')

def logout_page(request):
    logout(request)
    messages.success(request, "Vous êtes déconnecter")
    return redirect('index')

def register_page(request):
    if request.method == 'POST':
        username = request.POST.get("username")
        last_name = request.POST.get("last_name")
        first_name = request.POST.get("first_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        password2 = request.POST.get("password2")
        is_superuser = request.POST.get("is_superuser")
        if is_superuser is None:
            is_superuser=0
        new_user = User.objects.create_user(username=username,
                                            last_name=last_name,
                                            first_name=first_name,
                                            email=email,
                                            password=password,
                                            is_superuser=is_superuser)
        new_user.save()
        return redirect("index")

    else:
        messages.error(request, 'Formulaire invalide')
        return render(request, 'register.html')

def manager(request):

    managers=User.objects.all()
    return render(request, 'liste_user.html', {'managers':managers})

def get_manager_info(request, manager_id):
    utilisateur = User.objects.get(id=manager_id)
    try:
        user_info = {
            'username': utilisateur.username,
            'last_name': utilisateur.last_name,
            'first_name': utilisateur.first_name,
            'email': utilisateur.email,
            }
        return JsonResponse(user_info)
    except User.DoesNotExist:
        return JsonResponse({'error': 'Utilisateur non trouvé'}, status=404)
def updateManager(request):
    if request.POST.get('managerId'):
        username = request.POST.get("username")
        last_name = request.POST.get("last_name")
        first_name = request.POST.get("first_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        password2 = request.POST.get("password2")
        is_superuser = request.POST.get("is_superuser")
        if is_superuser is None:
            is_superuser = 0
        try:
            manager_id = request.POST.get('managerId')
            print(manager_id)
            utilisateur = User.objects.get(pk=manager_id)
            utilisateur.username = username
            utilisateur.last_name = last_name
            utilisateur.first_name = first_name
            utilisateur.email = email
            utilisateur.is_superuser = is_superuser
            if password:
                password = make_password(password)
                utilisateur.password = password
            utilisateur.save()
            messages.success(request, f"Modification effectuée avec succès")
        except Exception as e:
            messages.error(request, f"Modification non effectuée. {e}")
        return redirect("user_manager")

def removeManager(request):
    if request.POST.get('manager_id'):
        try:
            manager = User.objects.get(pk=request.POST.get('manager_id'))
            manager.delete()
            messages.success(request, f"L'utilisateur {manager.username} a été supprimer avec succès")
        except Exception as e:
            messages.error(request, f"Desoler nous avons rencontré une erreur {e}.")
    return redirect("user_manager")

def index(request):
    return render(request, 'login.html')

def dashboard(request):
    return render(request,'Dashboard.html')

def classes(request):
    if request.method == "POST":
        niveau_id = request.POST.get("niveau_id")
        classe = request.POST.get("classe")
        try:
            room = Classes.objects.create(classe_name=classe, niveau_id=niveau_id)
            room.save()
            messages.success(request, f"La classe de {classe} a été crée avec succès")
        except:
            messages.error(request, "Cette classe existe déjà.")
        return redirect("classes")

    else:
        niveaux = Niveau.objects.all()
        room = Classes.objects.all()
        return render(request,'classes.html', {'classes':room, 'niveaux':niveaux})

    
def removeClasse(request):
    if request.POST.get('classe_id'):
        try:
            classe = Classes.objects.get(pk=request.POST.get('classe_id'))
            classe.delete()
            messages.success(request, f"La classe de {classe.classe_name} a été supprimer avec succès")
        except Exception as e:
            messages.error(request, f"Desoler nous avons rencontré une erreur {e}.")
    return redirect("classes")

def createNiveau(request):
    if request.method == "POST":
        section = request.POST.get("section")
        niveau = Niveau.objects.create(name=section)
        niveau.save()
    return redirect("classes")

def typeAbonnements(req):
    if req.method == "POST":
        form = TypeAbonnementsForm(req.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(req, "Ajout de type d'abonnement avec succès")
            except Exception as e:
                messages.error(req, f"Cet abonnement existe déjà {e}")
            return redirect("typeAbonnements")
    else:
        form = TypeAbonnementsForm()
    return render(req, 'type-abonnements.html', { 'form': form , 'typeAbonnements': TypeAbonnements.objects.all()} )

def removeAbonnement(req):
    if req.POST.get('abonnement_id'):
        try:
            abonnement = TypeAbonnements.objects.get(pk=req.POST.get('abonnement_id'))
            abonnement.delete()
            messages.success(req, f"L'abonnement {abonnement.type} a été supprimé avec succès")
        except Exception as e:
            messages.error(req, f"Désoler nous avons rencontré une erreur {e}.")
    return redirect("typeAbonnements")

def editAbonnement(req, abonnement_id):
    abonnement = TypeAbonnements.objects.get(pk = abonnement_id)
    if req.method == 'POST':
        form = TypeAbonnementsForm(req.POST, instance=abonnement)
        if form.is_valid():
            form.save()
            return redirect("typeabonnements")
    else:
        form = TypeAbonnementsForm(instance=abonnement)
    return render(req, 'type-abonnements.html', { 'form': form , 'typeabonnements': TypeAbonnements.objects.all()} )

def calculer_jours_restants(personne):
    date_actuelle = timezone.now().date()
    date_expiration = personne.date_abonnement + timedelta(personne.type_abonnement.duree_jours)
    jours_restants = (date_expiration - date_actuelle).days
    if jours_restants == 0 :
        personne.is_abonne = 0
        personne.save()
    return jours_restants

def abonnement(req):
    teachersAndStudents = CustomUser.objects.all().filter(is_abonne = 0)
    typeAbonnements = TypeAbonnements.objects.all()
    return render(req, 'abonnement.html', {"typeAbonnements": typeAbonnements, "teachers": teachersAndStudents.filter(user_type = "Enseignant"), "students": teachersAndStudents.filter(user_type = "Eleve")})

def abonner(req, element_id):
    element = CustomUser.objects.get(pk = element_id)
    if req.method == 'POST':
        typeAbonnementId = req.POST.get("type_abonnement_id")
        element.is_abonne = True
        element.type_abonnement_id = typeAbonnementId
        element.date_abonnement = timezone.now().date()
        element.save()
    return redirect("abonnement")


def abonnes(req):
    teachersAndStudents = CustomUser.objects.all().filter(is_abonne = 1)
    liste_jours = {}
    for teachersAndStudent in teachersAndStudents:
        liste_jours[teachersAndStudent.id] = calculer_jours_restants(teachersAndStudent)
    return render(req, 'abonnes.html', {"teachers": teachersAndStudents.filter(user_type = "Enseignant"), "students": teachersAndStudents.filter(user_type = "Eleve"), "jours_restants": liste_jours})


def desabonner(request):
    if request.POST.get('teacher_id'):
        try:
            teacher = CustomUser.objects.get(pk=request.POST.get('teacher_id'))
            teacher.is_abonne = 0
            teacher.save()
            messages.success(request, f"L'abonné {teacher.firstname} a été supprimer avec succès")
        except Exception as e:
            messages.error(request, f"Desoler nous avons rencontré une erreur {e}.")
    if request.POST.get('student_id'):
        try:
            student = CustomUser.objects.get(pk=request.POST.get('student_id'))
            student.is_abonne = 0
            student.save()
            messages.success(request, f"L'abonné {student.firstname} a été supprimer avec succès")
        except Exception as e:
            messages.error(request, f"Desoler nous avons rencontré une erreur {e}.")
    return redirect("abonnes")


def inscrire(req):
    if req.method == 'POST':
        form = CustomUserForm(req.POST)
        if form.is_valid():
            form.save()
            return redirect("liste")
    else:

        form = CustomUserForm()
        return render(req, 'inscrire.html', {'form': form} )


def editCustomuser(req, user_id):
    user = CustomUser.objects.get(pk = user_id)
    if req.method == 'POST':
        form = CustomUserForm(req.POST, instance = user)
        if form.is_valid():
            form.save()
            return redirect("liste")
    else:
        form = CustomUserForm(instance = user)
    return render(req, 'inscrire.html', {'form': form})


def liste(request):
    teachersAndStudents = CustomUser.objects.all()

    return render(request, 'liste.html', {"teachers": teachersAndStudents.filter(user_type = "Enseignant"), "students": teachersAndStudents.filter(user_type = "Eleve")})


def removeStudent(req):
    if req.POST.get('student_id'):
        try:
            student = CustomUser.objects.get(pk=req.POST.get('student_id'))
            name = student.firstname
            student.delete()
            messages.success(req, f"L'élève {name} a été supprimé avec succès")
        except Exception as e:
            messages.error(req, f"Désoler nous avons rencontré une erreur {e}.")
    return redirect("liste")


def removeTeacher(req):
    if req.POST.get('teacher_id'):
        try:
            teacher = CustomUser.objects.get(pk=req.POST.get('teacher_id'))
            name = teacher.firstname
            teacher.delete()
            messages.success(req, f"L'élève {name} a été supprimé avec succès")
        except Exception as e:
            messages.error(req, f"Désoler nous avons rencontré une erreur {e}.")
    return redirect("liste")

def importer_fichier(request):
    if request.method == 'POST' and request.FILES['fichier']:
        fichier = request.FILES['fichier']

        try:
            if fichier.name.endswith('.xls'):
                # Utiliser xlrd pour les fichiers .xls
                book = xlrd.open_workbook(file_contents=fichier.read())
                sheet = book.sheet_by_index(0)
                # Enregistrement des données dans la base de données
                for row_num in range(1, sheet.nrows):  # Commencer à partir de la deuxième ligne
                    row = sheet.row_values(row_num)
                    instance = CustomUser(
                        lastname=row[0],
                        firstname=row[1],
                        genre=row[2],
                        user_type=row[3],
                        classe_id=row[4],
                    )
                    instance.save()
                messages.success(request, 'Données importées avec succès')
                return redirect("liste")
            elif fichier.name.endswith('.xlsx'):
                # Utiliser openpyxl pour les fichiers .xlsx
                workbook = load_workbook(fichier, read_only=True)
                sheet = workbook.active
                # Enregistrement des données dans la base de données
                for row in sheet.iter_rows(min_row=2, values_only=True):  # Commencer à partir de la deuxième ligne

                    instance = CustomUser(
                        lastname=row[0],
                        firstname=row[1],
                        genre=row[2],
                        user_type=row[3],
                        classe_id=row[4],
                    )
                    instance.save()
                messages.success(request, 'Données importées avec succès')
                return redirect("liste")
            else:
                raise ValueError('Format de fichier non pris en charge')
        except Exception as e:
            return JsonResponse({'erreur': str(e)}, status=500)

    return JsonResponse({'erreur': 'Requête incorrecte'}, status=400)
